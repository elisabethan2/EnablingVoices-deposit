#!/usr/bin/env python3
"""
Enabling Voices - compute_review_statistics.py
==============================================
Single-source computation of every screening/validation statistic reported in
the manuscript (Sections 3.2-3.5). Reads the deposited screening and annotation
outputs and writes two artefacts:

  - validation_metrics.xlsx : human-readable table of all reported statistics
  - validation_metrics.json : machine-readable version + run metadata

Re-running this script on the deposited inputs regenerates every number, so the
manuscript's figures all trace to one computation.

Determinism: all operations are deterministic counts/agreement statistics; no
sampling or stochastic step is involved, so no random seed applies.

Usage:
  python compute_review_statistics.py                  # writes to the default output dir
  python compute_review_statistics.py --output-dir DIR # write elsewhere
  python compute_review_statistics.py --overwrite      # replace existing outputs

Run from the repository root; input paths in the CONFIGURATION block are
repo-relative. Existing outputs are never replaced unless --overwrite is passed,
so a re-run cannot silently clobber the deposited copies of
validation_metrics.xlsx / .json (whose Run_info records the original run).
"""

import os
import re
import sys
import json
import argparse
import platform
from datetime import datetime, timezone

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# =============================================================================
# CONFIGURATION - input paths (repo-relative) and output directory
# =============================================================================
# Paths are relative to the repository root; run this script from there.
BASE = "data/screening"

ANNOTATOR_FILE  = f"{BASE}/master_comparison.xlsx"                              # annotator A/B + Llama + Qwen R1, 120-paper subset (sheet All120Papers)
COMPARISON_FILE = f"{BASE}/model_comparison_report.xlsx"                        # Llama vs Qwen R1 comparison report (sheet All_Comparisons)
QWEN_R2_FILE    = f"{BASE}/enabling_voices_qwen_v3_20260312_230707.xlsx"        # Qwen Round 2, full 470-record run (sheet Screening)

# Outputs are written here by default: a dedicated directory, NOT the working
# directory. Writing to the working directory would place the regenerated files
# on top of the deposited copies when the script is run from the repository root
# or from code/08_validation_metrics/, destroying the original run's Run_info.
DEFAULT_OUTPUT_DIR = "code/08_validation_metrics/outputs"

OUTPUT_BASENAMES = ("validation_metrics.json", "validation_metrics.xlsx")

# Provenance note: the final inclusion set and the four final-narrowing exclusions
# (below) are documented in enabling_voices_master_inclusion.xlsx (Master Inclusion
# Tracker) and Section 3.5; that file is deposited separately and is not read here.

# =============================================================================
# DOCUMENTED CONSTANTS - the final inclusion set and the human-validated set.
# These are stated explicitly (not inferred) and each carries its provenance.
# =============================================================================
# Final included set (N=6). The master_inclusion tracker lists 10, reflecting an
# earlier stage; the following four were excluded on final full-text review and
# are therefore NOT in the final set:
#   COV5183 Favela (C1/C4), COV11296 Rudzicz (C2), COV342 Faisal (C4), COV5639 Rass (C2/C3)
FINAL_INCLUDED = {"COV2660", "COV510", "COV5607", "COV5630", "COV5688", "COV1113"}
# Bailey, Sheehy, Obiorah, Purohit, Xygkou, Stara

# The 12 human-validated subset includes = 9 T1 consensus (derived below from the
# annotator data) PLUS the following 3, which were includes by tiebreaker /
# adjudication rather than by A==B consensus:
#   COV3004 Zhou and COV5169 Mowri (T2 tiebreaker: Qwen + the including annotator),
#   COV1113 Stara (PI full-text adjudication of an equal-confidence disagreement).
NON_CONSENSUS_VALIDATED = {"COV3004", "COV5169", "COV1113"}

FINAL_NARROWING_EXCLUSIONS = {"COV5183", "COV11296", "COV342", "COV5639"}

# =============================================================================
# HELPERS
# =============================================================================
def cov(x):
    """Extract the stable Covidence ID (e.g. 'COV5630') from a paper_id/filename."""
    m = re.match(r"(COV\d+)", str(x).strip(), flags=re.IGNORECASE)
    return m.group(1).upper() if m else str(x).strip()


def is_include(series):
    """Boolean 'include' mask, robust to True/False, 1/0, and 'include'/'exclude'."""
    def one(v):
        if isinstance(v, bool):
            return v
        s = str(v).strip().lower()
        if s in ("true", "1", "1.0", "yes", "y", "include", "included", "inc"):
            return True
        return s.startswith("inc")
    return series.map(one)


def has_decision(series):
    """True where a usable (non-failed, non-null) decision is present."""
    def one(v):
        if v is None:
            return False
        s = str(v).strip().lower()
        return s not in ("", "nan", "none", "fail", "failed", "error", "na")
    return series.map(one)


def cohen_kappa(a, b):
    """Cohen's kappa for two boolean arrays of equal length."""
    a = np.asarray(a, dtype=bool); b = np.asarray(b, dtype=bool)
    n = len(a)
    if n == 0:
        return float("nan")
    po = np.mean(a == b)
    p_a, p_b = a.mean(), b.mean()
    pe = p_a * p_b + (1 - p_a) * (1 - p_b)
    return float((po - pe) / (1 - pe)) if (1 - pe) != 0 else float("nan")


def binary_metrics(pred, ref):
    """Diagnostic-accuracy metrics treating `ref` (include=True) as the reference."""
    pred = np.asarray(pred, dtype=bool); ref = np.asarray(ref, dtype=bool)
    tp = int(np.sum(pred & ref)); fp = int(np.sum(pred & ~ref))
    tn = int(np.sum(~pred & ~ref)); fn = int(np.sum(~pred & ref))
    sens = tp / (tp + fn) if (tp + fn) else float("nan")
    spec = tn / (tn + fp) if (tn + fp) else float("nan")
    prec = tp / (tp + fp) if (tp + fp) else float("nan")
    acc = (tp + tn) / len(pred) if len(pred) else float("nan")
    return {"n": int(len(pred)), "TP": tp, "FP": fp, "TN": tn, "FN": fn,
            "sensitivity": round(sens, 3), "specificity": round(spec, 3),
            "precision": round(prec, 3), "accuracy": round(acc, 3),
            "kappa": round(cohen_kappa(pred, ref), 3)}


def require(path):
    if not os.path.exists(path):
        sys.exit(f"ERROR: input not found: {path}\n  Edit the CONFIGURATION block to match your paths.")
    return path

def parse_args(argv=None):
    ap = argparse.ArgumentParser(
        description="Recompute the reported screening and validation statistics "
                    "from the deposited screening and annotation outputs.")
    ap.add_argument("--output-dir", default=DEFAULT_OUTPUT_DIR, metavar="DIR",
                    help=f"where to write the outputs (default: {DEFAULT_OUTPUT_DIR})")
    ap.add_argument("--overwrite", action="store_true",
                    help="allow existing output files to be replaced")
    return ap.parse_args(argv)


def check_output_targets(out_dir, overwrite):
    """Refuse to replace existing outputs unless explicitly permitted.

    Checked before any computation, so a refusal never leaves one of the two
    output files rewritten and the other stale.
    """
    if overwrite:
        return
    existing = [os.path.join(out_dir, name) for name in OUTPUT_BASENAMES
                if os.path.exists(os.path.join(out_dir, name))]
    if existing:
        sys.exit(
            "ERROR: refusing to overwrite existing output file(s):\n"
            + "".join(f"    {path}\n" for path in existing)
            + "  If these are the deposited copies, replacing them destroys the\n"
              "  original run's Run_info record (timestamp, library versions, input\n"
              "  paths), which cannot be recovered from a re-run.\n"
              "  Write elsewhere with --output-dir DIR, or pass --overwrite if you\n"
              "  really do intend to replace them."
        )


# =============================================================================
# LOAD
# =============================================================================
def main(argv=None):
    args = parse_args(argv)
    out_dir = args.output_dir
    check_output_targets(out_dir, args.overwrite)

    for p in (ANNOTATOR_FILE, COMPARISON_FILE, QWEN_R2_FILE):
        require(p)

    ann = pd.read_excel(ANNOTATOR_FILE, sheet_name="All120Papers")
    ann["cov"] = ann["paper_id"].map(cov)
    a_inc = is_include(ann["decision_A"]); b_inc = is_include(ann["decision_B"])
    llama_inc = is_include(ann["llama_decision"]); qwenR1_inc = is_include(ann["qwen_decision"])

    consensus_inc_mask = a_inc & b_inc
    consensus_exc_mask = (~a_inc) & (~b_inc)
    disagree_mask = a_inc ^ b_inc
    consensus_mask = consensus_inc_mask | consensus_exc_mask        # A==B
    consensus_ref = consensus_inc_mask                              # reference 'include' on consensus papers

    subset_covs = set(ann["cov"])
    consensus_inc_covs = set(ann.loc[consensus_inc_mask, "cov"])
    human_validated_12 = consensus_inc_covs | NON_CONSENSUS_VALIDATED

    # Qwen Round 2 (full corpus)
    q2 = pd.read_excel(QWEN_R2_FILE, sheet_name="Screening")
    q2["cov"] = q2["_filename"].map(cov)
    q2_inc_covs = set(q2.loc[is_include(q2["include"]), "cov"])

    # Llama vs Qwen R1 comparison report (for the valid-only agreement denominator)
    cmp = pd.read_excel(COMPARISON_FILE, sheet_name="All_Comparisons")
    cmp_valid = cmp[has_decision(cmp["include_llama"]) & has_decision(cmp["include_qwen"])]

    R = {}  # results

    # --- 1. Validation-subset composition --------------------------------------
    R["subset_composition"] = {
        "n_subset": int(len(ann)),
        "consensus_include": int(consensus_inc_mask.sum()),
        "consensus_exclude": int(consensus_exc_mask.sum()),
        "disagreement": int(disagree_mask.sum()),
        "consensus_total_AeqB": int(consensus_mask.sum()),
    }

    # --- 2. Inter-annotator agreement (A vs B, all 120) ------------------------
    R["inter_annotator"] = {
        "n": int(len(ann)),
        "raw_agreement_pct": round(100 * float((a_inc == b_inc).mean()), 1),
        "kappa": round(cohen_kappa(a_inc, b_inc), 3),
    }

    # --- 3. Model vs human-consensus reference (consensus papers only) ---------
    cons = ann[consensus_mask]
    R["llama_vs_consensus"] = binary_metrics(is_include(cons["llama_decision"]),
                                             is_include(cons["decision_A"]) & is_include(cons["decision_B"]))
    R["qwenR1_vs_consensus"] = binary_metrics(is_include(cons["qwen_decision"]),
                                              is_include(cons["decision_A"]) & is_include(cons["decision_B"]))

    # --- 4. Llama vs Qwen R1 agreement (two denominators) ----------------------
    R["llama_qwenR1_agreement"] = {
        "all_120_raw_pct": round(100 * float((llama_inc == qwenR1_inc).mean()), 1),
        "all_120_kappa": round(cohen_kappa(llama_inc, qwenR1_inc), 3),
        "all_120_n": int(len(ann)),
        "valid_only_raw_pct": round(100 * float(
            (is_include(cmp_valid["include_llama"]) == is_include(cmp_valid["include_qwen"])).mean()), 1),
        "valid_only_n": int(len(cmp_valid)),
        "note": "valid_only excludes records with a failed Llama text-extraction; all_120 counts failures as exclude.",
    }

    # --- 5. Retrospective comparison against the final included set (N=6) -------
    final_in_subset = FINAL_INCLUDED & subset_covs
    llama_inc_covs = set(ann.loc[llama_inc, "cov"])
    qwenR1_inc_covs = set(ann.loc[qwenR1_inc, "cov"])
    q2_inc_in_subset = q2_inc_covs & subset_covs
    R["retrospective_vs_final6"] = {
        "reference": "final included set (N=6); descriptive/post-hoc, not a diagnostic gold standard",
        "n_final": len(FINAL_INCLUDED),
        "n_final_in_subset": len(final_in_subset),
        "llama_R1": {"includes_total": len(llama_inc_covs),
                     "final_recovered": len(llama_inc_covs & FINAL_INCLUDED),
                     "over_inclusions": len(llama_inc_covs) - len(llama_inc_covs & FINAL_INCLUDED)},
        "qwen_R1": {"includes_total": len(qwenR1_inc_covs),
                    "final_recovered": len(qwenR1_inc_covs & FINAL_INCLUDED),
                    "over_inclusions": len(qwenR1_inc_covs) - len(qwenR1_inc_covs & FINAL_INCLUDED),
                    "missed_final": len(FINAL_INCLUDED & subset_covs) - len(qwenR1_inc_covs & FINAL_INCLUDED)},
        "qwen_R2_in_subset": {"includes_in_subset": len(q2_inc_in_subset),
                              "final_recovered": len(q2_inc_in_subset & FINAL_INCLUDED),
                              "over_inclusions": len(q2_inc_in_subset) - len(q2_inc_in_subset & FINAL_INCLUDED)},
    }

    # --- 6. Qwen Round 2 include set, split against the validation subset -------
    tp = q2_inc_covs & human_validated_12
    fp = (q2_inc_covs & subset_covs) - human_validated_12
    new = q2_inc_covs - subset_covs
    R["qwenR2_include_split"] = {
        "total_includes": len(q2_inc_covs),
        "overlap_human_validated": len(tp),
        "human_reviewed_exclusions": len(fp),
        "new_outside_subset": len(new),
        "sum_check": len(tp) + len(fp) + len(new),
    }

    # --- 7. Inclusion accounting (12 human-validated -> final 6) ---------------
    removed = human_validated_12 - FINAL_INCLUDED
    R["inclusion_accounting"] = {
        "human_validated_includes": len(human_validated_12),
        "final_included": len(FINAL_INCLUDED),
        "removed_at_full_text": len(removed),
        "removed_covs": sorted(removed),
        "final_all_from_subset": FINAL_INCLUDED.issubset(human_validated_12),
        "new_papers_reviewed": len(new),
        "new_papers_included": len(new & FINAL_INCLUDED),
    }

    # Integrity checks
    R["checks"] = {
        "subset_parts_sum_to_n": (R["subset_composition"]["consensus_include"]
                                  + R["subset_composition"]["consensus_exclude"]
                                  + R["subset_composition"]["disagreement"]) == len(ann),
        "human_validated_is_12": len(human_validated_12) == 12,
        "qwenR2_split_sums_to_total": R["qwenR2_include_split"]["sum_check"] == len(q2_inc_covs),
        "twelve_minus_removed_is_six": (len(human_validated_12) - len(removed)) == len(FINAL_INCLUDED),
    }

    meta = {
        "generated_utc": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
        "script": os.path.basename(__file__),
        "determinism": "deterministic; no random seed applies (no sampling/stochastic steps)",
        "python": platform.python_version(),
        "pandas": pd.__version__, "numpy": np.__version__,
        "inputs": {"annotator_file": ANNOTATOR_FILE, "comparison_file": COMPARISON_FILE,
                   "qwen_r2_file": QWEN_R2_FILE},
        "notes": "Reference standards are stated per statistic; consensus = A==B annotator agreement.",
    }

    os.makedirs(out_dir, exist_ok=True)
    write_json(R, meta, out_dir)
    write_xlsx(R, meta, out_dir)
    print_summary(R)
    return R


# =============================================================================
# OUTPUT WRITERS
# =============================================================================
def write_json(R, meta, out_dir):
    out = os.path.join(out_dir, "validation_metrics.json")
    with open(out, "w", encoding="utf-8") as f:
        json.dump({"metadata": meta, "metrics": R}, f, indent=2, ensure_ascii=False)
    print(f"  wrote {out}")


def write_xlsx(R, meta, out_dir):
    wb = Workbook()
    HEAD = Font(name="Arial", bold=True, color="FFFFFF")
    BODY = Font(name="Arial")
    FILL = PatternFill("solid", fgColor="003D73")  # SDU navy
    wrap = Alignment(wrap_text=True, vertical="top")

    def style_header(ws, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(1, c); cell.font = HEAD; cell.fill = FILL
        ws.freeze_panes = "A2"

    def autosize(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # Sheet 1: flat metrics table
    ws = wb.active; ws.title = "Metrics"
    cols = ["Section", "Statistic", "Value", "n", "Reference standard", "Notes"]
    ws.append(cols)
    rows = [
        ("Subset composition", "Consensus include", R["subset_composition"]["consensus_include"], R["subset_composition"]["n_subset"], "Annotator A==B", ""),
        ("Subset composition", "Consensus exclude", R["subset_composition"]["consensus_exclude"], R["subset_composition"]["n_subset"], "Annotator A==B", ""),
        ("Subset composition", "Disagreement (A!=B)", R["subset_composition"]["disagreement"], R["subset_composition"]["n_subset"], "Annotator A vs B", "PI-adjudicated"),
        ("Inter-annotator", "Raw agreement (%)", R["inter_annotator"]["raw_agreement_pct"], R["inter_annotator"]["n"], "Annotator A vs B", ""),
        ("Inter-annotator", "Cohen's kappa", R["inter_annotator"]["kappa"], R["inter_annotator"]["n"], "Annotator A vs B", ""),
        ("Llama vs consensus", "Sensitivity", R["llama_vs_consensus"]["sensitivity"], R["llama_vs_consensus"]["n"], "Human consensus", "include-enriched subset"),
        ("Llama vs consensus", "Specificity", R["llama_vs_consensus"]["specificity"], R["llama_vs_consensus"]["n"], "Human consensus", ""),
        ("Llama vs consensus", "Precision", R["llama_vs_consensus"]["precision"], R["llama_vs_consensus"]["n"], "Human consensus", ""),
        ("Llama vs consensus", "Accuracy", R["llama_vs_consensus"]["accuracy"], R["llama_vs_consensus"]["n"], "Human consensus", ""),
        ("Llama vs consensus", "Cohen's kappa", R["llama_vs_consensus"]["kappa"], R["llama_vs_consensus"]["n"], "Human consensus", ""),
        ("Qwen R1 vs consensus", "Sensitivity", R["qwenR1_vs_consensus"]["sensitivity"], R["qwenR1_vs_consensus"]["n"], "Human consensus", "include-enriched subset"),
        ("Qwen R1 vs consensus", "Specificity", R["qwenR1_vs_consensus"]["specificity"], R["qwenR1_vs_consensus"]["n"], "Human consensus", ""),
        ("Qwen R1 vs consensus", "Precision", R["qwenR1_vs_consensus"]["precision"], R["qwenR1_vs_consensus"]["n"], "Human consensus", ""),
        ("Qwen R1 vs consensus", "Accuracy", R["qwenR1_vs_consensus"]["accuracy"], R["qwenR1_vs_consensus"]["n"], "Human consensus", ""),
        ("Qwen R1 vs consensus", "Cohen's kappa", R["qwenR1_vs_consensus"]["kappa"], R["qwenR1_vs_consensus"]["n"], "Human consensus", ""),
        ("Llama vs Qwen R1", "Raw agreement (%) - all 120", R["llama_qwenR1_agreement"]["all_120_raw_pct"], R["llama_qwenR1_agreement"]["all_120_n"], "Model vs model", "failures counted as exclude"),
        ("Llama vs Qwen R1", "Cohen's kappa - all 120", R["llama_qwenR1_agreement"]["all_120_kappa"], R["llama_qwenR1_agreement"]["all_120_n"], "Model vs model", ""),
        ("Llama vs Qwen R1", "Raw agreement (%) - valid only", R["llama_qwenR1_agreement"]["valid_only_raw_pct"], R["llama_qwenR1_agreement"]["valid_only_n"], "Model vs model", "excludes 9 failed extractions"),
        ("Retrospective (final-6)", "Llama R1 final recovered / over-inclusions", f"{R['retrospective_vs_final6']['llama_R1']['final_recovered']} / {R['retrospective_vs_final6']['llama_R1']['over_inclusions']}", R["retrospective_vs_final6"]["llama_R1"]["includes_total"], "Final included set (N=6)", "descriptive, post-hoc"),
        ("Retrospective (final-6)", "Qwen R1 final recovered / over / missed", f"{R['retrospective_vs_final6']['qwen_R1']['final_recovered']} / {R['retrospective_vs_final6']['qwen_R1']['over_inclusions']} / {R['retrospective_vs_final6']['qwen_R1']['missed_final']}", R["retrospective_vs_final6"]["qwen_R1"]["includes_total"], "Final included set (N=6)", "descriptive, post-hoc"),
        ("Retrospective (final-6)", "Qwen R2 final recovered / over (in-subset)", f"{R['retrospective_vs_final6']['qwen_R2_in_subset']['final_recovered']} / {R['retrospective_vs_final6']['qwen_R2_in_subset']['over_inclusions']}", R["retrospective_vs_final6"]["qwen_R2_in_subset"]["includes_in_subset"], "Final included set (N=6)", "within validation subset"),
        ("Qwen R2 include split", "Overlap human-validated (TP)", R["qwenR2_include_split"]["overlap_human_validated"], R["qwenR2_include_split"]["total_includes"], "12 human-validated includes", ""),
        ("Qwen R2 include split", "Human-reviewed exclusions (FP)", R["qwenR2_include_split"]["human_reviewed_exclusions"], R["qwenR2_include_split"]["total_includes"], "validation subset", ""),
        ("Qwen R2 include split", "New (outside subset)", R["qwenR2_include_split"]["new_outside_subset"], R["qwenR2_include_split"]["total_includes"], "n/a", ""),
        ("Inclusion accounting", "Human-validated includes", R["inclusion_accounting"]["human_validated_includes"], "", "9 consensus + 2 tiebreaker + 1 adjudication", ""),
        ("Inclusion accounting", "Removed at full-text", R["inclusion_accounting"]["removed_at_full_text"], "", "PI full-text review", ", ".join(R["inclusion_accounting"]["removed_covs"])),
        ("Inclusion accounting", "Final included", R["inclusion_accounting"]["final_included"], "", "finalised criteria", "all from validation subset"),
    ]
    for r in rows:
        ws.append(list(r))
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.font = BODY; c.alignment = wrap
    style_header(ws, len(cols))
    autosize(ws, [22, 36, 14, 8, 34, 30])

    # Sheet 2: confusion matrices
    ws2 = wb.create_sheet("Confusion_matrices")
    cols2 = ["Comparison", "n", "TP", "FP", "TN", "FN", "Sensitivity", "Specificity", "Precision", "Accuracy", "Kappa"]
    ws2.append(cols2)
    for name, key in [("Llama vs consensus", "llama_vs_consensus"), ("Qwen R1 vs consensus", "qwenR1_vs_consensus")]:
        m = R[key]
        ws2.append([name, m["n"], m["TP"], m["FP"], m["TN"], m["FN"],
                    m["sensitivity"], m["specificity"], m["precision"], m["accuracy"], m["kappa"]])
    for row in ws2.iter_rows(min_row=2):
        for c in row:
            c.font = BODY
    style_header(ws2, len(cols2))
    autosize(ws2, [22, 6, 6, 6, 6, 6, 12, 12, 11, 10, 8])

    # Sheet 3: run info
    ws3 = wb.create_sheet("Run_info")
    ws3.append(["Field", "Value"])
    for k, v in meta.items():
        ws3.append([k, json.dumps(v) if isinstance(v, dict) else str(v)])
    for row in ws3.iter_rows(min_row=2):
        for c in row:
            c.font = BODY; c.alignment = wrap
    style_header(ws3, 2)
    autosize(ws3, [22, 90])

    out = os.path.join(out_dir, "validation_metrics.xlsx")
    wb.save(out)
    print(f"  wrote {out}")


def print_summary(R):
    print("\n" + "=" * 60)
    print("VALIDATION METRICS - summary")
    print("=" * 60)
    sc = R["subset_composition"]
    print(f"  Subset: {sc['consensus_include']} consensus-include / "
          f"{sc['consensus_exclude']} consensus-exclude / {sc['disagreement']} disagree (n={sc['n_subset']})")
    ia = R["inter_annotator"]
    print(f"  Inter-annotator: {ia['raw_agreement_pct']}%  kappa={ia['kappa']}")
    lc, qc = R["llama_vs_consensus"], R["qwenR1_vs_consensus"]
    print(f"  Llama vs consensus:  sens={lc['sensitivity']} spec={lc['specificity']} prec={lc['precision']} acc={lc['accuracy']} k={lc['kappa']}  (TP/FP/TN/FN={lc['TP']}/{lc['FP']}/{lc['TN']}/{lc['FN']})")
    print(f"  Qwen R1 vs consensus: sens={qc['sensitivity']} spec={qc['specificity']} prec={qc['precision']} acc={qc['accuracy']} k={qc['kappa']}  (TP/FP/TN/FN={qc['TP']}/{qc['FP']}/{qc['TN']}/{qc['FN']})")
    lq = R["llama_qwenR1_agreement"]
    print(f"  Llama-Qwen R1: all-120 {lq['all_120_raw_pct']}% (k={lq['all_120_kappa']}) | valid-only {lq['valid_only_raw_pct']}% (n={lq['valid_only_n']})")
    rt = R["retrospective_vs_final6"]
    print(f"  Retrospective vs final-6: Llama {rt['llama_R1']['final_recovered']}/6 (+{rt['llama_R1']['over_inclusions']} FP) | "
          f"Qwen R1 {rt['qwen_R1']['final_recovered']}/6 (+{rt['qwen_R1']['over_inclusions']} FP, {rt['qwen_R1']['missed_final']} FN) | "
          f"Qwen R2 {rt['qwen_R2_in_subset']['final_recovered']}/6 (+{rt['qwen_R2_in_subset']['over_inclusions']} FP in-subset)")
    sp = R["qwenR2_include_split"]
    print(f"  Qwen R2 split: {sp['overlap_human_validated']} + {sp['human_reviewed_exclusions']} + {sp['new_outside_subset']} = {sp['sum_check']} (of {sp['total_includes']})")
    ic = R["inclusion_accounting"]
    print(f"  Inclusion: {ic['human_validated_includes']} validated - {ic['removed_at_full_text']} removed = {ic['final_included']} final; all from subset: {ic['final_all_from_subset']}")
    print(f"  Integrity checks: {R['checks']}")
    print("=" * 60)


if __name__ == "__main__":
    main()
