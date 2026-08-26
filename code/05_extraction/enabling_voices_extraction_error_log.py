#!/usr/bin/env python3
"""
Enabling Voices — Full-Text Extraction Error Log (generator)
============================================================
Documents discrepancies between the trialled LLM full-text extraction
(Llama 3.1 8B, 14 March 2026; code/05_extraction/enabling_voices_extraction_fulltext.py)
and the study characteristics charted MANUALLY for Section 4. The extraction was
trialled and NOT used for the reported results (paper Section 3.6; Supplementary D.8);
this script regenerates the error record so the cautionary finding is auditable.

Works on either workbook:
  - the curated 6-paper file (friendly headers: "N (cond.)", "AI Components", ...)
  - the raw output  enabling_voices_extraction_fulltextWholePaper_*.xlsx
    (dotted headers: "population.n_with_condition", "ai_technology.ai_components", ...;
     setting lives in the Methodology sheet)

Usage : python enabling_voices_extraction_error_log.py <workbook.xlsx> [output_dir]
        output_dir defaults to the folder containing the workbook (persistent storage).
Output: extraction_error_log.md  and  extraction_errors.csv
"""
import sys, csv, os
import openpyxl

XLSX = sys.argv[1] if len(sys.argv) > 1 else "enabling_voices_final_6papers.xlsx"
OUTDIR = sys.argv[2] if len(sys.argv) > 2 else (os.path.dirname(os.path.abspath(XLSX)) or ".")
os.makedirs(OUTDIR, exist_ok=True)

# ── Verified Section 4 / Table 4 values (manually charted reference) ──────────
GT = {
    "COV2660": dict(study="Bailey 2026",  model="GPT-3.5", n=3,
                    setting="university laboratory (US)", uses_gpt4=False, setting_ok=True),
    "COV5607": dict(study="Obiorah 2021", model="rule-based AAC; image-captioning APIs, no LLM",
                    n=11, setting="multi-site: clinic, restaurants, community",
                    uses_gpt4=False, setting_ok="partial"),
    "COV5630": dict(study="Purohit 2023", model="GPT-3.5 (ChatGPT)", n=8,
                    setting="secondary analysis of AphasiaBank transcripts; no live deployment",
                    uses_gpt4=False, setting_ok=False),
    "COV510":  dict(study="Sheehy 2024",  model="GPT-3.5 (Stage 2)", n=20,
                    setting="long-term care facility (Canada)", uses_gpt4=False, setting_ok=True),
    "COV1113": dict(study="Stara 2021",   model="rule-based (Anne; ASR+TTS), no LLM", n=20,
                    setting="home (Ancona, Italy)", uses_gpt4=False, setting_ok=True),
    "COV5688": dict(study="Xygkou 2024",  model="GPT-4", n=8,
                    setting="home (UK)", uses_gpt4=True, setting_ok=True),
}

# ── Read sheets, keying rows by whichever COV-ID column exists ────────────────
COV_KEYS = ("COV ID", "_cov_id", "COV_ID")
wb = openpyxl.load_workbook(XLSX, data_only=True)
def sheet_rows(name):
    if name not in wb.sheetnames: return {}
    ws = wb[name]; H = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    cov_col = next((k for k in COV_KEYS if k in H), None)
    out = {}
    if cov_col is None: return out
    ci = H.index(cov_col)
    for r in range(2, ws.max_row + 1):
        cov = ws.cell(r, ci + 1).value
        if cov is None: continue
        out[str(cov).strip()] = {H[c-1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
    return out

ov, ait, meth = sheet_rows("Overview"), sheet_rows("AI_Technology"), sheet_rows("Methodology")

def pick(row, *names):
    for n in names:
        if n in row and row[n] not in (None, ""): return row[n]
    return None

def has_gpt4(s): return s is not None and ("gpt-4" in str(s).lower() or "gpt4" in str(s).lower())

rows = []
for cov, g in GT.items():
    o, a, m = ov.get(cov, {}), ait.get(cov, {}), meth.get(cov, {})
    comps = pick(a, "AI Components", "ai_technology.ai_components")
    ex_n  = pick(o, "N (cond.)", "population.n_with_condition")
    conf  = pick(o, "Confidence", "extraction_quality.overall_confidence")
    ex_set = (pick(o, "Setting") or pick(m, "setting.setting_type", "Setting", "setting.setting_detail"))

    if has_gpt4(comps) and not g["uses_gpt4"]:
        rows.append([cov, g["study"], "AI model", str(comps), g["model"],
                     "Model misattribution (spurious GPT-4)", conf])
    elif g["uses_gpt4"] and has_gpt4(comps):
        rows.append([cov, g["study"], "AI model", str(comps), g["model"],
                     "Correct (GPT-4 genuinely used)", conf])
    try:
        if int(ex_n) != int(g["n"]):
            rows.append([cov, g["study"], "N (participants)", ex_n, g["n"], "Participant miscount", conf])
    except (TypeError, ValueError):
        rows.append([cov, g["study"], "N (participants)", ex_n, g["n"], "Participant value unparseable", conf])
    if g["setting_ok"] is False:
        rows.append([cov, g["study"], "Setting", ex_set, g["setting"], "Setting misidentified", conf])
    elif g["setting_ok"] == "partial":
        rows.append([cov, g["study"], "Setting", ex_set, g["setting"],
                     "Setting flattened (multi-site -> single)", conf])

# ── CSV ──────────────────────────────────────────────────────────────────────
csv_path = os.path.join(OUTDIR, "extraction_errors.csv")
with open(csv_path, "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(["COV_ID","Study","Field","Extracted_value","Verified_value","Error_type","Model_confidence"])
    w.writerows(rows)

# ── Markdown ─────────────────────────────────────────────────────────────────
gpt4 = sorted({r[1] for r in rows if "spurious GPT-4" in r[5]})
nerr = sorted({r[1] for r in rows if r[2] == "N (participants)" and "miscount" in r[5]})
serr = sorted({r[1] for r in rows if r[2] == "Setting"})
def cell(x): return str(x).replace("|", " · ").replace("\n", " ").strip()
md = []
md.append("# Full-text extraction — error log\n")
md.append("**Status.** Discrepancies between the trialled LLM full-text extraction (Llama 3.1 8B, "
          "14 March 2026) and the study characteristics charted **manually** for Section 4. The "
          "extraction was trialled and **not used** for the reported results; deposited only to "
          "document its failure modes (paper §3.6; Supplementary D.8). Verified values are the "
          "manually charted Section 4 / Table 4 characteristics; extracted values are read directly "
          "from the deposited workbook. The run processed 10 candidates as they stood in March 2026; "
          "this log covers the six studies in the final included set.\n")
md.append("## Summary\n")
md.append(f"- **Model misattribution (spurious GPT-4):** {', '.join(gpt4)}. Only Xygkou 2024 genuinely used GPT-4.")
md.append(f"- **Participant miscount:** {', '.join(nerr) or 'none'}.")
md.append(f"- **Setting misidentified / flattened:** {', '.join(serr) or 'none'}.\n")
md.append("## Per-field discrepancies\n")
md.append("| COV ID | Study | Field | Extracted | Verified (Section 4) | Error type | Model confidence |")
md.append("|---|---|---|---|---|---|---|")
for r in rows: md.append("| " + " | ".join(cell(x) for x in r) + " |")
md.append("\n## Confidence pattern\n")
md.append("Model-reported confidence did not track correctness. The two highest-confidence extractions "
          "(Sheehy 2024 and Purohit 2023, both confidence 4) are the two that carry a spurious GPT-4, "
          "while the one study that genuinely used GPT-4 (Xygkou 2024) carries the lowest confidence (2). "
          "Confidence was therefore highest where the extraction was wrong.\n")
md.append("## Mechanism\n")
md.append("The errors follow from the windowed extraction design. A memory-limited GPU (Tesla V100, 32 GB) "
          "could not hold the full schema and a long article in one pass, so each paper was split into "
          "overlapping token windows, the full schema run on each, then merged. List fields (AI components) "
          "merge by **union**, so a single window mentioning GPT-4 — in related work, a citation, or a "
          "hallucination — propagates to the final field; whole-document aggregates (Obiorah's N = 11 across "
          "phases) cannot be recovered because no single window holds them; scalar fields take the first "
          "non-empty value, so an early window's setting guess (Purohit's 'university lab') overrides later, "
          "more accurate windows. Confidence is the minimum across windows, computed independently of the "
          "union, which is why wrong fields ride through at moderate confidence.\n")
open(os.path.join(OUTDIR, "extraction_error_log.md"), "w", encoding="utf-8").write("\n".join(md))
print(f"Wrote to {OUTDIR}:  extraction_error_log.md  +  extraction_errors.csv")
print(f"Flagged {len(rows)} field-level entries across {len(set(r[1] for r in rows if 'Correct' not in r[5]))} studies with errors.")
