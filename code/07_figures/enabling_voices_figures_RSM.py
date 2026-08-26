#!/usr/bin/env python3
"""
Enabling Voices — Figure generation for Research Synthesis Methods submission
=============================================================================

Regenerates the four figures used in the manuscript (all in Section 3,
the LLM-assisted screening methodology), in a form that meets the
Cambridge/RSM artwork requirements:

  * NO titles baked into the image  -> captions live in the manuscript text
    (Cambridge: "Have captions supplied at the end of the manuscript text,
     instead of as part of the figure file.")
  * Vector PDF output (resolution-independent; accepted format alongside
    EPS/TIFF) PLUS a 600 dpi PNG for embedding in the working Word draft.
    Cambridge requires >=600 dpi for combination figures if raster is used;
    vector PDF sidesteps the dpi question entirely.
  * Consecutive numbering matching the manuscript (Fig1..Fig4).
  * Optional filename prefix for the Cambridge convention
    "[First-author surname]Fig1" (set AUTHOR_PREFIX below before final upload).

Figures (manuscript numbering):
    Fig1 — Llama 3.1 8B confidence-score distribution by decision category
    Fig2 — Llama vs. Qwen Round-1 disagreement areas (by criterion)
    Fig3 — PRISMA-ScR flow diagram (combined pipeline + human validation)
    Fig4 — Retrospective agreement matrices (4 panels) vs. adjudicated ground truth

DATA PROVENANCE (every number is read from a file; nothing is hard-coded):
    Fig1 : ROUND1_XLSX  ['Summary'] : confidence x _category          -> n = 461
    Fig2 : COMPARE_XLSX ['Disagreements'] : per-criterion field diffs  -> n = 19
    Fig3 : static flow values (no data file; encoded from the PRISMA log)
    Fig4 : MASTER_XLSX  ['All120Papers'] : paper_id, llama_decision, qwen_decision
           QWEN_R2_XLSX ['Screening']    : _filename, include
           GROUND_TRUTH : the six final-included COV records

NOTE on the Round-1 input: Figure 1 plots the 461 records that received a valid Round-1 decision (470 submitted − 9 text-extraction failures = 28 include + 408 exclude + 25 manual review). ROUND1_XLSX must point to the corrected workbook that yields this count.

DETERMINISM: these figures are fully deterministic. No sampling, shuffling,
or random splits are involved, so no random seed is required. Re-running on
identical inputs reproduces byte-stable counts.

Environment (record alongside your run log):
    Python 3.10+, matplotlib >= 3.5, openpyxl >= 3.0, numpy >= 1.21
    pip install matplotlib openpyxl numpy --break-system-packages

Run:
    python enabling_voices_figures_RSM.py
"""

import os
import datetime
from collections import Counter

import numpy as np
import openpyxl
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.patches import FancyBboxPatch

# ════════════════════════════════════════════════════════════════════════════
# CONFIGURATION  — set these three things before running on UCloud
# ════════════════════════════════════════════════════════════════════════════
# 1) Where the input workbooks live (persistent storage on UCloud, NOT /tmp):
# Root folder that holds outputs/. Defaults to the UCloud run location;
# override with the ENABLING_VOICES_DIR env var to reproduce elsewhere.
BASE_DIR   = os.environ.get("ENABLING_VOICES_DIR", "/work/EnablingPapers150126")
INPUT_DIR  = os.path.join(BASE_DIR, "outputs")
# 2) Where figures are written (persistent storage):
OUTPUT_DIR = os.path.join(BASE_DIR, "outputs", "figures_RSM")
# 3) Filename prefix for the Cambridge convention "[Surname]Fig1".
#    Leave "" during drafting; set to the lead-author surname before upload,
#    e.g. AUTHOR_PREFIX = "Andersen"  ->  AndersenFig1.pdf
AUTHOR_PREFIX = "Andersen"

# Output formats. PDF = vector submission file; PNG = preview / Word embed.
SAVE_PDF = True
SAVE_PNG = True
PNG_DPI  = 600          # >=600 satisfies Cambridge for combination raster art

# ── Input filenames (relative to INPUT_DIR) ────────────────────────────────
ROUND1_XLSX  = "enabling_voices_round1_V2_20260206_170842.xlsx"   # # corrected; yields n = 461
COMPARE_XLSX = "model_comparison_report.xlsx"
MASTER_XLSX  = "adjudication_output/master_comparison.xlsx"
QWEN_R2_XLSX ="qwen_screening_v3_120326_3/enabling_voices_qwen_v3_20260312_230707.xlsx" 

# ── The six final-included studies (Fig4 ground truth) ──────────────────────
GROUND_TRUTH = {
    "COV2660_Bailey2025.pdf",
    "COV510_Sheehy2024.pdf",
    "COV5607_Obiorah2021.pdf",
    "COV5688_Xygkou2024.pdf",
    "COV5630_Purohit2023.pdf",
    "COV1113_Stara2021.pdf",
}

# ════════════════════════════════════════════════════════════════════════════
# SDU palette  (from SDU_SLIDE_INSTRUCTIONS.md)
# ════════════════════════════════════════════════════════════════════════════
SDU_NAVY  = "#003D73"   # primary / dementia / agreement-include
SDU_BLUE  = "#0077C8"   # accent / agreement-exclude
SDU_MID   = "#4DA3D4"   # gradient midpoint
SDU_LIGHT = "#E0EEF7"   # light fill / off-diagonal cells
GRAY      = "#666666"
DARK      = "#333333"
WHITE     = "#FFFFFF"

COL_INCLUDE = SDU_BLUE
COL_EXCLUDE = SDU_NAVY
COL_MANUAL  = SDU_MID

os.makedirs(OUTPUT_DIR, exist_ok=True)


# ── Saving helper: one figure -> PDF (+ PNG), titles already omitted ────────
def _save(fig, n):
    stem = f"{AUTHOR_PREFIX}Fig{n}"
    written = []
    if SAVE_PDF:
        p = os.path.join(OUTPUT_DIR, stem + ".pdf")
        fig.savefig(p, bbox_inches="tight", facecolor=WHITE, pad_inches=0.08)
        written.append(p)
    if SAVE_PNG:
        p = os.path.join(OUTPUT_DIR, stem + ".png")
        fig.savefig(p, dpi=PNG_DPI, bbox_inches="tight",
                    facecolor=WHITE, pad_inches=0.08)
        written.append(p)
    plt.close(fig)
    for p in written:
        print(f"  saved {os.path.basename(p)}  ({os.path.getsize(p)//1024} KB)")


def _path(fn):
    return os.path.join(INPUT_DIR, fn)


# ════════════════════════════════════════════════════════════════════════════
# FIGURE 1 — Llama confidence-score distribution by decision category
# ════════════════════════════════════════════════════════════════════════════
def fig1_llama_confidence():
    print("Figure 1: Llama confidence distribution")
    wb = openpyxl.load_workbook(_path(ROUND1_XLSX))
    ws = wb["Summary"]
    hdr = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    ci = hdr.index("confidence")
    cat = hdr.index("_category")

    data = {"Include": Counter(), "Exclude": Counter(), "Manual Review": Counter()}
    n_valid = 0
    for r in range(2, ws.max_row + 1):
        c = ws.cell(r, cat + 1).value
        v = ws.cell(r, ci + 1).value
        if c in data and v is not None:
            data[c][int(v)] += 1
            n_valid += 1

    scores = [1, 2, 3, 4, 5]
    exc = [data["Exclude"].get(s, 0) for s in scores]
    inc = [data["Include"].get(s, 0) for s in scores]
    man = [data["Manual Review"].get(s, 0) for s in scores]
    print(f"  n = {n_valid}  Exclude={exc} Include={inc} Manual={man}")

    fig, ax = plt.subplots(figsize=(9, 5.5))
    fig.patch.set_facecolor(WHITE)
    ax.set_facecolor(WHITE)

    x = np.arange(len(scores))
    w = 0.26
    b1 = ax.bar(x - w, exc, w, label="Exclude", color=COL_EXCLUDE, zorder=3)
    b2 = ax.bar(x,     inc, w, label="Include", color=COL_INCLUDE, zorder=3)
    b3 = ax.bar(x + w, man, w, label="Manual Review", color=COL_MANUAL, zorder=3)

    for bars in (b1, b2, b3):
        for bar in bars:
            h = bar.get_height()
            if h > 0:
                ax.text(bar.get_x() + bar.get_width() / 2, h + 3, str(int(h)),
                        ha="center", va="bottom", fontsize=8, color=DARK)

    ax.set_xlabel("Confidence Score (1 = very uncertain, 5 = very confident)",
                  fontsize=11, color=DARK, labelpad=8)
    ax.set_ylabel("Number of Papers", fontsize=11, color=DARK, labelpad=8)
    # NO title — caption is supplied in the manuscript text.

    ax.set_xticks(x)
    ax.set_xticklabels([str(s) for s in scores], fontsize=11, color=DARK)
    ax.tick_params(axis="y", labelsize=10, colors=DARK)
    ax.tick_params(axis="x", colors=DARK)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_color("#CCCCCC")
    ax.spines["bottom"].set_color("#CCCCCC")
    ax.yaxis.grid(True, linestyle="--", alpha=0.5, color="#CCCCCC", zorder=0)
    ax.set_axisbelow(True)

    leg = ax.legend(fontsize=10, framealpha=0.9, edgecolor="#CCCCCC",
                    loc="upper left", ncol=1)
    for t in leg.get_texts():
        t.set_color(DARK)

    plt.tight_layout()
    _save(fig, 1)


# ════════════════════════════════════════════════════════════════════════════
# FIGURE 2 — Llama vs. Qwen Round-1 disagreement areas (by criterion)
# ════════════════════════════════════════════════════════════════════════════
def fig2_disagreement_areas():
    print("Figure 2: disagreement areas")
    wb = openpyxl.load_workbook(_path(COMPARE_XLSX))
    ws = wb["Disagreements"]
    hdr = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = [{hdr[c - 1]: ws.cell(r, c).value
             for c in range(1, ws.max_column + 1)}
            for r in range(2, ws.max_row + 1)]
    n = len(rows)

    # Per-criterion counts (a paper can differ on more than one criterion).
    # C1 = population, C2 = AI technology, C3 = communication support.
    c2 = sum(1 for d in rows
             if d.get("has_ai_llama") != d.get("has_ai_qwen"))
    c3 = sum(1 for d in rows
             if d.get("has_communication_support_llama")
             != d.get("has_communication_support_qwen"))
    c1 = sum(1 for d in rows
             if (d.get("has_dementia_llama") != d.get("has_dementia_qwen")
                 or d.get("has_aphasia_llama") != d.get("has_aphasia_qwen")))
    print(f"  n = {n}  C2(AI)={c2} C3(comm)={c3} C1(pop)={c1}")

    # bottom -> top order to match the manuscript figure
    labels = ["AI technology boundary (C2)",
              "Communication focus (C3)",
              "Population scope (C1)"]
    counts = [c2, c3, c1]
    colours = [SDU_NAVY, SDU_BLUE, SDU_MID]

    fig, ax = plt.subplots(figsize=(9.5, 4.2))
    fig.patch.set_facecolor(WHITE)
    ax.set_facecolor(WHITE)

    bars = ax.barh(labels, counts, color=colours, height=0.55,
                   zorder=3, edgecolor=WHITE, linewidth=1.5)
    for bar, cnt in zip(bars, counts):
        ax.text(bar.get_width() + 0.3, bar.get_y() + bar.get_height() / 2,
                f"{cnt}/{n}  ({100 * cnt / n:.0f}%)",
                va="center", ha="left", fontsize=11, color=DARK)

    ax.set_xlim(0, max(counts) + 6)
    ax.set_xlabel("Number of disagreement cases\n"
                  "(papers may contribute to multiple categories)",
                  fontsize=10, color=DARK, labelpad=8)
    # NO title — caption in manuscript.
    ax.tick_params(axis="y", labelsize=11, colors=DARK)
    ax.tick_params(axis="x", labelsize=10, colors=DARK)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_visible(False)
    ax.spines["bottom"].set_color("#CCCCCC")
    ax.xaxis.grid(True, linestyle="--", alpha=0.4, color="#CCCCCC", zorder=0)
    ax.set_axisbelow(True)

    plt.tight_layout()
    _save(fig, 2)


# ════════════════════════════════════════════════════════════════════════════
# FIGURE 3 — PRISMA-ScR flow diagram  (dynamic box heights; no overflow)
# ════════════════════════════════════════════════════════════════════════════
def fig3_prisma():
    print("Figure 3: PRISMA-ScR flow diagram")

    NAVY, LIGHT = SDU_NAVY, "#D6EAF8"
    AMBER_B, AMBER_F, AMBER_T = "#B07800", "#FFF3CD", "#5C3D00"

    FIG_W = 16
    MAIN_X, MAIN_W = 7.8, 7.0
    SIDE_X, SIDE_W = 14.25, 3.2
    LABEL_X, LABEL_W = 0.85, 1.5
    MAIN_R = MAIN_X + MAIN_W / 2
    SIDE_L = SIDE_X - SIDE_W / 2
    GAP = 0.42
    FS_MAIN, FS_SIDE, FS_PHASE = 12, 11, 9
    LH_MAIN = 12 / 72 * 1.30
    LH_SIDE = 11 / 72 * 1.30
    PAD = 0.30

    def box_h(lines, lh):
        return len(lines) * lh + PAD

    ITEMS = [
        (['Records identified via database searching  (n = 7,147)',
          'Scopus, ComDisDome, LLBA & MLA, combined (n = 2,113)',
          'Web of Science (1,780)   ·   MEDLINE (1,298)   ·   PsycINFO (543)',
          'CINAHL (488)   ·   IEEE Xplore (374)   ·   ACM Digital Library (263)',
          'Google Scholar (200)   ·   Sociological Abstracts (88)'],
         {0}, LIGHT, None, None, 'IDENTIFICATION'),
        (['2,495 duplicate records removed by Covidence',
          '4,652 records screened (title/abstract in Covidence)'],
         None, LIGHT, None, None, None),
        (['Covidence abstract screening',
          '4,174 excluded  →  478 records assessed for eligibility'],
         None, LIGHT,
         ['Excluded at Covidence (n = 4,174)', 'Not relevant to population,',
          'technology, or communication', 'scope (title/abstract level)'],
         {0}, 'SCREENING'),
        (['478 records imported to Covidence',
          '8 without PDFs: manual review  →  all 8 excluded',
          '470 records submitted to LLM screening'],
         None, LIGHT,
         ['Excluded (n = 8)', 'Review article; conference poster;',
          'non-English; not relevant (×4);', 'practitioner magazine article'],
         {0}, None),
        (['Llama 3.1 8B — Round 1 screening  (n = 470)',
          '28 include   |   408 exclude   |   25 manual review',
          '9 text-extraction failures → human-reviewed, all excluded'],
         {0}, LIGHT,
         ['Excluded by Llama (n = 408)', '25 flagged → validation subset',
          '9 text-extraction failures', '→ human-reviewed, all excluded'],
         {0}, 'LLM\nROUND 1'),
        (['120-paper stratified validation subset',
          '(28 Llama includes + 25 manual review + 67 sampled excludes)'],
         None, LIGHT, None, None, None),
        (['Human double-annotation  (8 annotators, 30 papers each)',
          'Qwen 2.5 7B Round 1 validation on same 120 papers',
          '→ 9 T1 consensus  +  2 T2 tiebreaker  +  1 PI adjudicated  =  12 includes'],
         {0}, LIGHT,
         ['Human consensus excludes', 'n = 85 of 94 consensus papers',
          '26 disagreement cases', '→ Tier 2 adjudication',
          '3 ADJUDICATE → PI decision'], {0}, 'HUMAN\nVALIDATION'),
        (['Group criteria refinement',
          'Informed by human–LLM disagreement analysis',
          '→ Finalised 4-criterion definitions + Qwen Round 2 prompt'],
         None, LIGHT, None, None, 'CRITERIA\nREFINEMENT'),
        (['Qwen 2.5 7B Round 2 — full corpus  (n = 470)',
          '41 includes (38 Tier A + 3 Tier B)   |   278 flagged   |   151 exclude',
          '8 also human includes   |   13 also human excludes   |   20 new for PI review'],
         {0}, LIGHT,
         ['Of 41 Qwen Round 2 includes:', '8 also human includes',
          '13 also human excludes (auto-excl.)', '20 new papers → PI review',
          '(8 + 13 + 20 = 41)'], {0}, 'LLM\nROUND 2'),
        (['PI review — two tracks',
          'Track A: 12 human includes',
          '→ −3 (Altwala 2024: C4; Yuan 2023: C3; Favela 2023: C1 & C4)',
          '→ −3 discrepancy check (Zhou 2022: C2-WoZ;',
          '   Mowri 2025: C1; Du 2024: C2 & C3);',
          '   Obiorah 2021 retained (C2)  =  6 retained',
          'Track B: 20 new Qwen Round 2 papers  →  0 included'],
         {0}, LIGHT,
         ['Track A excluded (n = 6)', '−3: Altwala 2024 (C4)',
          '     Yuan 2023 (C3)', '     Favela 2023 (C1 & C4)',
          '−3: Zhou 2022 (C2-WoZ)', '     Mowri 2025 (C1)',
          '     Du 2024 (C2 & C3)', 'Track B: 20/20 new excluded'], {0}, 'PI\nREVIEW'),
        (['FINAL INCLUDED STUDIES  (n = 6)',
          'Bailey 2026  ·  Sheehy 2024  ·  Obiorah 2021',
          'Xygkou 2024  ·  Purohit 2023  ·  Stara 2021'],
         {0}, NAVY, None, None, 'INCLUSION'),
    ]

    heights = [box_h(it[0], LH_MAIN) for it in ITEMS]
    total_h = sum(heights) + GAP * (len(ITEMS) - 1)
    FIG_H = total_h + 0.3   # small margin only — no title row

    boxes = []
    top = total_h
    for (ml, bi, fill, sl, sb, ph), mh in zip(ITEMS, heights):
        cy = top - mh / 2
        boxes.append((mh, cy, ml, bi, fill, sl, sb, ph))
        top -= (mh + GAP)

    fig, ax = plt.subplots(figsize=(FIG_W, FIG_H))
    fig.patch.set_facecolor(WHITE)
    ax.set_facecolor(WHITE)
    ax.set_xlim(0, FIG_W)
    ax.set_ylim(0, FIG_H)
    ax.axis("off")

    def mbox(cx, cy, w, h, lines, bold_idx=None, fill=LIGHT, fs=FS_MAIN):
        bold_idx = bold_idx or set()
        ax.add_patch(FancyBboxPatch((cx - w / 2, cy - h / 2), w, h,
                     boxstyle="round,pad=0.08", facecolor=fill,
                     edgecolor=NAVY, linewidth=1.8, zorder=3))
        tcol = WHITE if fill == NAVY else DARK
        n = len(lines)
        topL = cy + (n - 1) / 2 * LH_MAIN
        for i, line in enumerate(lines):
            ax.text(cx, topL - i * LH_MAIN, line, ha="center", va="center",
                    fontsize=fs, fontweight="bold" if i in bold_idx else "normal",
                    color=tcol, zorder=4, linespacing=1.1)

    def sbox(cy, h, lines, bold_idx=None, fs=FS_SIDE):
        bold_idx = bold_idx or set()
        ax.add_patch(FancyBboxPatch((SIDE_X - SIDE_W / 2, cy - h / 2), SIDE_W, h,
                     boxstyle="round,pad=0.08", facecolor=AMBER_F,
                     edgecolor=AMBER_B, linewidth=2.5, zorder=3))
        n = len(lines)
        topL = cy + (n - 1) / 2 * LH_SIDE
        for i, line in enumerate(lines):
            ax.text(SIDE_X, topL - i * LH_SIDE, line, ha="center", va="center",
                    fontsize=fs, fontweight="bold" if i in bold_idx else "normal",
                    color=AMBER_T, zorder=4, linespacing=1.1)

    def phase(cy, h, text):
        ax.add_patch(FancyBboxPatch((LABEL_X - LABEL_W / 2, cy - h / 2),
                     LABEL_W, h, boxstyle="round,pad=0.06", facecolor=NAVY,
                     edgecolor="none", zorder=3))
        ax.text(LABEL_X, cy, text, ha="center", va="center", fontsize=FS_PHASE,
                fontweight="bold", color=WHITE, zorder=4,
                multialignment="center")

    def darrow(x, y1, y2):
        ax.annotate("", xy=(x, y2 + 0.04), xytext=(x, y1 - 0.04),
                    arrowprops=dict(arrowstyle="->", color=NAVY, lw=2.0,
                                    mutation_scale=18), zorder=5)

    def sarrow(y):
        ax.annotate("", xy=(SIDE_L - 0.04, y), xytext=(MAIN_R + 0.04, y),
                    arrowprops=dict(arrowstyle="->", color=AMBER_B, lw=1.8,
                                    mutation_scale=14), zorder=5)

    prev_bottom = None
    for mh, cy, ml, bi, fill, sl, sb, ph in boxes:
        mbox(MAIN_X, cy, MAIN_W, mh, ml, bold_idx=bi, fill=fill)
        if ph:
            phase(cy, mh, ph)
        if sl:
            sh = box_h(sl, LH_SIDE)
            sbox(cy, sh, sl, bold_idx=sb)
            sarrow(cy)
        if prev_bottom is not None:
            darrow(MAIN_X, prev_bottom, cy + mh / 2)
        prev_bottom = cy - mh / 2

    plt.subplots_adjust(top=0.99, bottom=0.01, left=0.01, right=0.99)
    _save(fig, 3)


# ════════════════════════════════════════════════════════════════════════════
# FIGURE 4 — Retrospective agreement matrices (4 panels) vs. ground truth
# ════════════════════════════════════════════════════════════════════════════
def _is_inc(v):
    return str(v).strip().upper() == "INCLUDE" or str(v).strip().upper() == "TRUE"


def fig4_agreement_matrices():
    print("Figure 4: retrospective agreement matrices")

    # ── Load the 120-paper master comparison ────────────────────────────────
    wb = openpyxl.load_workbook(_path(MASTER_XLSX))
    ws = wb["All120Papers"]
    H = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    pid_i = H.index("paper_id")
    lla_i = H.index("llama_decision")
    qw1_i = H.index("qwen_decision")
    master = []
    for r in range(2, ws.max_row + 1):
        pid = ws.cell(r, pid_i + 1).value
        if pid is None:
            continue
        master.append({
            "pid": str(pid).strip(),
            "llama": _is_inc(ws.cell(r, lla_i + 1).value),
            "qwen1": _is_inc(ws.cell(r, qw1_i + 1).value),
        })
    ids120 = [m["pid"] for m in master]
    n120 = len(master)

    # ── Load Qwen Round-2 full-corpus decisions, restrict to the 120 ─────────
    wb2 = openpyxl.load_workbook(_path(QWEN_R2_XLSX), read_only=True)
    ws2 = wb2["Screening"]
    H2 = [ws2.cell(1, c).value for c in range(1, ws2.max_column + 1)]
    fi = H2.index("_filename")
    ii = H2.index("include")
    q2 = {}
    for r in range(2, ws2.max_row + 1):
        fn = ws2.cell(r, fi + 1).value
        if fn is None:
            continue
        q2[os.path.basename(str(fn).strip())] = bool(ws2.cell(r, ii + 1).value)
    wb2.close()
    for m in master:
        m["qwen2"] = q2.get(os.path.basename(m["pid"]), False)

    # ── Build the four 2x2 count matrices ────────────────────────────────────
    def matrix(pred_key, ref_key):
        # rows: predicted Include/Exclude ; cols: reference Include/Exclude
        c = Counter()
        for m in master:
            pred = m[pred_key]
            ref = (m["pid"] in GROUND_TRUTH) if ref_key == "gt" else m[ref_key]
            c[(pred, ref)] += 1
        # [TP, FP, FN, TN]
        return c[(True, True)], c[(True, False)], c[(False, True)], c[(False, False)]

    panels = [
        ("Llama 3.1 8B",          "Adjudicated\nGround Truth", matrix("llama", "gt"), True),
        ("Qwen 2.5 7B\nRound 1",  "Adjudicated\nGround Truth", matrix("qwen1", "gt"), True),
        ("Qwen 2.5 7B\nRound 2",  "Adjudicated\nGround Truth", matrix("qwen2", "gt"), True),
        ("Qwen 2.5 7B\nRound 1",  "Llama\n3.1 8B",             matrix("qwen1", "llama"), False),
    ]
    for row_lab, _, (tp, fp, fn, tn), gt in panels:
        agree = (tp + tn) / n120
        print(f"  {row_lab.replace(chr(10), ' '):24s} "
              f"TP={tp} FP={fp} FN={fn} TN={tn}  agree={agree:.1%}")

    # ── Draw ─────────────────────────────────────────────────────────────────
    fig, axes = plt.subplots(1, 4, figsize=(15, 4.6))
    fig.patch.set_facecolor(WHITE)

    for ax, (row_lab, col_lab, (tp, fp, fn, tn), is_gt) in zip(axes, panels):
        ax.set_facecolor(WHITE)
        ax.set_xlim(0, 2)
        ax.set_ylim(0, 2.7)
        ax.axis("off")
        total = tp + fp + fn + tn
        agree = (tp + tn) / total

        # cell grid: rows top=Include, bottom=Exclude ; cols left=Include, right=Exclude
        cells = {(0, 0): tp, (0, 1): fp, (1, 0): fn, (1, 1): tn}
        fills = {(0, 0): SDU_NAVY, (1, 1): SDU_BLUE,
                 (0, 1): SDU_LIGHT, (1, 0): SDU_LIGHT}
        for (ri, cidx), val in cells.items():
            bg = fills[(ri, cidx)]
            ax.add_patch(FancyBboxPatch(
                (cidx + 0.04, 1 - ri + 0.04), 0.92, 0.92,
                boxstyle="round,pad=0.02", facecolor=bg,
                edgecolor=WHITE, linewidth=2, zorder=2))
            tcol = WHITE if bg in (SDU_NAVY, SDU_BLUE) else SDU_NAVY
            cx, cy = cidx + 0.5, 1 - ri + 0.58
            ax.text(cx, cy, str(val), ha="center", va="center",
                    fontsize=24, fontweight="bold", color=tcol, zorder=3)
            ax.text(cx, cy - 0.30, f"({100 * val / total:.0f}%)",
                    ha="center", va="center", fontsize=10, color=tcol, zorder=3)

        # column labels (reference)
        ax.text(0.5, 2.06, "Include", ha="center", va="bottom",
                fontsize=10, color=DARK)
        ax.text(1.5, 2.06, "Exclude", ha="center", va="bottom",
                fontsize=10, color=DARK)
        ax.text(1.0, 2.28, col_lab, ha="center", va="bottom",
                fontsize=10.5, fontweight="bold", color=SDU_NAVY,
                multialignment="center")

        # row labels (predicted)
        ax.text(-0.04, 1.5, "Include", ha="right", va="center",
                fontsize=10, color=DARK)
        ax.text(-0.04, 0.5, "Exclude", ha="right", va="center",
                fontsize=10, color=DARK)
        ax.text(-0.42, 1.0, row_lab, ha="center", va="center",
                rotation=90, fontsize=10.5, fontweight="bold",
                color=SDU_NAVY, multialignment="center")

        # agreement line above the panel
        if is_gt:
            sub = f"Agreement: {agree:.1%}  (n = {total})"
        else:
            sub = f"Agreement: {agree:.1%}  (n = {total})\nModel-only comparison (no ground truth)"
        ax.text(1.0, 2.62, sub, ha="center", va="bottom",
                fontsize=9, color=GRAY, multialignment="center")

    plt.subplots_adjust(left=0.05, right=0.99, top=0.86, bottom=0.02, wspace=0.55)
    _save(fig, 4)


# ════════════════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("=" * 64)
    print("Enabling Voices — RSM figure generation")
    print("run:", datetime.datetime.now().isoformat(timespec="seconds"))
    print("input dir :", INPUT_DIR)
    print("output dir:", OUTPUT_DIR)
    print("=" * 64)

    fig1_llama_confidence()
    fig2_disagreement_areas()
    fig3_prisma()
    fig4_agreement_matrices()

    print("-" * 64)
    print("Done. Files in", OUTPUT_DIR)
    for f in sorted(os.listdir(OUTPUT_DIR)):
        if f.endswith((".pdf", ".png")):
            print(f"  {f}  ({os.path.getsize(os.path.join(OUTPUT_DIR, f))//1024} KB)")
