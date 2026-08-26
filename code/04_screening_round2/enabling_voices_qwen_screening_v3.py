#!/usr/bin/env python3
"""
Enabling Voices – Qwen Screening v3
====================================
Screens papers using the criteria finalised in the group definitions document
(March 2026), running Qwen 2.5-7B-Instruct on the pool of papers that Llama
excluded in Round 1 and that have NOT been human-annotated.

This serves as a methodological safety net: Llama is inclusive and has few
false negatives, but this run gives an auditable second opinion and catches any
borderline cases Llama may have missed.

The prompt implements:
  C1 – Population   : dementia (incl. PPA) or aphasia, confirmed diagnosis
  C2 – AI Technology: Tier A (explicit) + Tier B (implicit/functional)
                      per European Commission (2018) definition
  C3 – Communication: direct communicative support, not therapy training,
                      not detection, not indirect psychosocial only
  C4 – Study type   : empirical only (review articles auto-excluded)

Each criterion gets its own evidence field + confidence score (1–5).
Final include/exclude is derived from criterion-level outputs, not a single
black-box judgement.

Output: timestamped .xlsx in OUTPUT_DIR with one row per paper, all criterion
fields, and a FLAG_REVIEW column for papers needing human verification.

Usage:
  python enabling_voices_qwen_screening_v3.py

Background (UCloud):
  nohup python enabling_voices_qwen_screening_v3.py > qwen_v3.log 2>&1 &
  tail -f qwen_v3.log

Edit the CONFIGURATION section below to match your UCloud paths.
"""

import os
import sys
import json
import re
import gc
from pathlib import Path
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

# =============================================================================
# CONFIGURATION – update these paths for your UCloud setup
# =============================================================================

PDF_FOLDER   = "/work/EnablingPapers150126"
OUTPUT_DIR   = "/work/EnablingPapers150126/outputs/qwen_screening_v3"
OUTPUT_PREFIX = "enabling_voices_qwen_v3"

# Llama round-1 results file (used to build the target paper list)
LLAMA_FILE  = "/work/EnablingPapers150126/outputs/enabling_voices_round1_V2_20260206_170842.xlsx"

# Human-annotated 120-paper set (to EXCLUDE from this run – already processed)
QWEN_120_FILE = "/work/EnablingPapers150126/outputs/qwen_validation_20260206_222701.xlsx"

# Model
MODEL_NAME   = "Qwen/Qwen2.5-7B-Instruct"
TEMPERATURE  = 0.1
MAX_TOKENS_INPUT  = 1500   # chars of PDF text passed to the model (~1500 tokens)
MAX_TOKENS_OUTPUT = 1200   # max new tokens in model response

# Processing controls
TEST_MODE_LIMIT  = None    # Set to e.g. 5 for a quick test run
CHECKPOINT_EVERY = 10      # Save checkpoint every N papers

# HuggingFace token (if needed for gated models)
HF_TOKEN = None   # Replace with your token string if required

# =============================================================================
# PAPER LIST
# =============================================================================
# RUN_ALL_PAPERS = True  → run on all 470 papers Llama processed (recommended).
#   Gives a single coherent validation run with the finalised definitions,
#   replacing the patchwork of earlier prompts. ~1 GPU-hour at 8000 tokens.
#
# RUN_ALL_PAPERS = False → run only on the 350 Llama-excluded papers that were
#   not in the 120-paper human annotation set (original safety-net mode).
#
# PAPER_OVERRIDE → ignore both flags; use this explicit list instead.

RUN_ALL_PAPERS = True

PAPER_OVERRIDE = None   # e.g. ["COV123_Smith2020.pdf", ...]

# =============================================================================
# SCREENING PROMPT
# =============================================================================

SCREENING_PROMPT = """\
You are screening a research paper for a scoping review. The review is titled
"Enabling Voices" and examines how artificial intelligence (AI) technologies
support communication for people with dementia and/or aphasia.

Your task is to assess the paper against four criteria and return a structured
JSON response. Read the article text carefully before answering.

═══════════════════════════════════════════════════════════════
CRITERION 1 – TARGET POPULATION
═══════════════════════════════════════════════════════════════
INCLUDE if the study involves people with a CONFIRMED clinical diagnosis of:
  • Dementia (any subtype: Alzheimer's disease, vascular dementia, Lewy body
    dementia, frontotemporal dementia, etc.)
  • Primary Progressive Aphasia (PPA) – count as both dementia and aphasia
  • Aphasia (acquired: post-stroke aphasia, non-progressive aphasia, etc.)

EXCLUDE if the study involves ONLY:
  • People with suspected or probable dementia (not confirmed)
  • Mild Cognitive Impairment (MCI) without a dementia diagnosis
  • Healthy older adults or at-risk populations
  • Caregivers or family members as the primary study population
  • Other communication disorders (autism, cerebral palsy, developmental disorders)

INCLUDE (with flag) if the study has a MIXED sample that includes people with
dementia or aphasia alongside others, AND the paper reports outcomes specifically
for the target group, OR the technology is designed primarily for their use.

═══════════════════════════════════════════════════════════════
CRITERION 2 – AI TECHNOLOGY
═══════════════════════════════════════════════════════════════
Definition adopted for this review (European Commission, 2018):
  "AI refers to systems that display intelligent behaviour by analysing their
  environment and taking action – with some degree of autonomy – to achieve
  specific goals."

This includes: machine learning, deep learning, neural networks, natural
language processing (NLP), automatic speech recognition (ASR), large language
models (LLMs, e.g. GPT-4, ChatGPT), dialogue systems, computer vision for
understanding, word prediction using language models, and social robots with
adaptive/learning dialogue capabilities.

TWO-TIER CLASSIFICATION:

  TIER A – Explicit AI: The paper uses AI terminology explicitly in relation
  to the communication technology. Keywords: artificial intelligence, AI,
  machine learning, deep learning, neural network, LLM, GPT, NLP, natural
  language processing, ASR, automatic speech recognition, computer vision,
  dialogue system, conversational AI, word prediction, transformer, BERT.
  → Include directly on AI criterion. Extract the exact term(s) found.

  TIER B – Implicit/Functional AI: No explicit AI terminology, but the
  technology behaves according to the EC definition: it perceives user input,
  processes it with some degree of autonomy beyond fixed rule-based lookup,
  and generates contextually variable responses. Examples: social robots with
  "speech recognition and dialogue capabilities"; AAC devices with "predictive
  text"; virtual agents that "respond adaptively to voice commands".
  → Flag for human verification (do not auto-include).

EXCLUDE on AI criterion if:
  • The technology is a basic digital tool: pre-programmed content, fixed
    scripts, simple touchscreen with pre-stored phrases, video calling software,
    digital photo albums, basic reminder apps with fixed schedules
  • AI is mentioned only in the background or future work, not as a feature
    of the technology studied in the paper
  • A social robot is described ONLY by brand name with no AI claim — the
    paper must explicitly describe AI, NLP, or adaptive/learning behaviour

THE FOLLOWING ARE NOT AI TERMS — do not list them as explicit_ai_terms or
use them alone to justify tier=A_explicit:
  Interface/hardware: touchscreen, touch screen, tablet, head-worn display,
    Google Glass, computer-based system, desktop computer, mobile app, screen
  Multimedia/content: hypermedia, multimedia, pre-recorded audio, storyboard,
    interactive icons, photo album, slideshow, digital stories
  Communication platforms: Skype, Zoom, FaceTime, video calling, telephone
  Design methodology: user-centered design, iterative design, participatory
    design, co-design, assistive technology (as a category)
  Generic digital infrastructure: cloud computing, programmed alert system,
    Bluetooth, WiFi, sensors (without ML processing)
  Robot names without AI claims: PARO, NAO (without NLP claim), Pepper
    (without NLP claim), teleoperated robot, wizard-of-oz robot
  AAC devices without generation/prediction: picture symbols, pre-stored
    phrases, symbol boards, static vocabulary sets

WHAT DOES COUNT as Tier A AI (explicit terms to look for):
  artificial intelligence, AI, machine learning, deep learning, neural network,
  large language model, LLM, GPT, ChatGPT, BERT, transformer model,
  natural language processing, NLP, natural language understanding,
  automatic speech recognition, ASR (when used for understanding/response),
  dialogue system, dialogue management, conversational agent (with NLP),
  word prediction (ML-based), emotion recognition (ML-based),
  computer vision (for understanding, not just display),
  adaptive/personalised responses (must describe the learning mechanism)

IMPORTANT: Do NOT require AI to be the sole component. Include if an AI
component plays a meaningful role in enabling communicative interaction.

═══════════════════════════════════════════════════════════════
CRITERION 3 – COMMUNICATION SUPPORT
═══════════════════════════════════════════════════════════════
INCLUDE if communication is a PRIMARY OUTCOME or PRIMARY MECHANISM:
  • Human–machine communication: the AI system is a conversational partner
    (chatbot, social robot with dialogue, voice assistant, ECA)
  • Human–human communication mediated by AI: AAC systems, real-time
    speech-to-text, AI-generated conversation prompts facilitating human
    interaction
  • Communication scaffolding: the AI helps the person initiate, sustain,
    repair, or augment communicative interaction

EXCLUDE if the primary focus is ONLY:
  • Indirect psychosocial outcomes: loneliness, depression, anxiety, general
    wellbeing – where communication is not a studied mechanism or outcome
  • Cognitive training/stimulation: memory exercises, quiz games, arithmetic
    tasks, cognitive rehabilitation exercises (these are not communication)
  • Clinical assessment or detection: using AI to diagnose or screen for
    dementia or aphasia, or to assess language impairment
  • Functional daily living (ADL) support: medication reminders, appointment
    management, safety monitoring, GPS tracking
  • SPEECH AND LANGUAGE THERAPY TRAINING APPS: structured exercises designed
    to rehabilitate language functions through drills (e.g. naming therapy apps
    that use ASR to score word-retrieval exercises). Rationale: these target
    communication-as-rehabilitation-target, not communication-as-practice.
    EXAMPLE OF EXCLUSION: "aphaDIGITAL – a mobile app for speech therapy
    with ASR-based feedback on naming exercises" → EXCLUDE on C3.
    EXAMPLE OF INCLUSION: "a voice assistant that helps people with aphasia
    communicate their daily needs to caregivers" → INCLUDE on C3.
  • COGNITIVE TRAINING OR STIMULATION: activities designed to exercise
    cognitive functions (memory games, quizzes, arithmetic, naming exercises,
    categorisation tasks) even when delivered by a robot or AI system. A robot
    leading structured cognitive stimulation sessions is doing therapy, not
    communication support. Ask: is the AI being a CONVERSATION PARTNER, or
    is it being a DRILL INSTRUCTOR?
    EXAMPLE OF EXCLUSION: "Pepper robot conducts cognitive and socio-cognitive
    training sessions with structured tasks and verbal instructions" → EXCLUDE.
    EXAMPLE OF INCLUSION: "Pepper robot engages residents in open-ended
    conversation and responds to their topics of interest" → INCLUDE.

BORDERLINE – flag for human review if:
  • Study has both detection AND communication support aspects
  • Social robot whose function mixes cognitive stimulation with conversation
  • Usability study of a system designed for communication (include unless the
    system itself fails C2 or C3)

═══════════════════════════════════════════════════════════════
CRITERION 4 – STUDY TYPE
═══════════════════════════════════════════════════════════════
EXCLUDE: review articles, systematic reviews, scoping reviews, meta-analyses,
theoretical or conceptual papers without empirical data.
INCLUDE: empirical studies of any design (experimental, observational,
qualitative, mixed methods, case study, pilot, feasibility, design study,
co-design, RCT).

═══════════════════════════════════════════════════════════════
ARTICLE TEXT:
{article_text}
═══════════════════════════════════════════════════════════════

Return ONLY a valid JSON object. No preamble, no explanation outside the JSON.

{{
  "C4_study_type": {{
    "is_review_article": true_or_false,
    "study_type": "empirical / review / theoretical / unclear",
    "evidence": "Brief quote or description supporting the study type classification"
  }},

  "C1_population": {{
    "has_dementia": true_or_false_or_null,
    "has_aphasia": true_or_false_or_null,
    "has_ppa": true_or_false_or_null,
    "diagnosis_confirmed": true_or_false_or_null,
    "population_notes": "Who are the participants? Quote any diagnosis terms used.",
    "confidence": 1_to_5
  }},

  "C2_ai_technology": {{
    "tier": "A_explicit / B_implicit / none",
    "explicit_ai_terms": ["list of AI terms found verbatim, e.g. 'machine learning', 'GPT-4'"],
    "implicit_ai_indicators": ["functional indicators if Tier B, e.g. 'adaptive dialogue', 'speech recognition that responds'"],
    "technology_name": "Name of the system/device studied",
    "ai_evidence_quote": "Best direct quote from paper showing AI use (max 200 chars)",
    "is_basic_digital_tool": true_or_false,
    "is_ai_only_in_background": true_or_false,
    "confidence": 1_to_5
  }},

  "C3_communication_support": {{
    "is_direct_communication_support": true_or_false_or_null,
    "is_therapy_training_app": true_or_false,
    "is_detection_assessment_only": true_or_false,
    "is_indirect_psychosocial_only": true_or_false,
    "is_cognitive_training_only": true_or_false,
    "communication_role": "Description of how the AI supports (or fails to support) communication",
    "confidence": 1_to_5
  }},

  "screening_decision": {{
    "include": true_or_false,
    "flag_for_human_review": true_or_false,
    "exclude_reason": "C1 / C2 / C3 / C4 / none (list all that apply, e.g. 'C2,C3')",
    "overall_confidence": 1_to_5,
    "decision_rationale": "Concise explanation of the decision (max 150 chars)"
  }}
}}

CONFIDENCE SCALE (use for each criterion AND overall):
  1 = Abstract only or very little text; major uncertainty
  2 = Some information but key details missing
  3 = Enough to decide but one or more criteria are ambiguous
  4 = Good evidence; decision is clear
  5 = Very clear; explicit statements for all relevant criteria

CRITICAL CONSISTENCY RULE — read before setting include:
  include = true  ONLY IF ALL of the following hold:
    • C4: is_review_article = false
    • C1: has_dementia OR has_aphasia OR has_ppa = true, AND diagnosis_confirmed = true
    • C2: tier = "A_explicit" OR "B_implicit"
          AND is_basic_digital_tool = false
          AND is_ai_only_in_background = false
    • C3: is_direct_communication_support = true
          AND is_therapy_training_app = false
          AND is_detection_assessment_only = false

  If C2.tier = "none" → include MUST be false. A paper with no AI technology
  cannot be included even if population and communication criteria are met.
  If C2.tier = "B_implicit" → set flag_for_human_review = true.

FLAG FOR HUMAN REVIEW if any of these apply:
  • Overall confidence ≤ 2
  • C2 tier = B_implicit (Tier B AI requires human verification)
  • C3 is borderline (detection AND communication aspects both present)
  • Any criterion confidence ≤ 2 on a paper you would otherwise include
"""

# =============================================================================
# UTILITIES
# =============================================================================

def load_paper_list(llama_path, qwen120_path, run_all=True):
    """
    Build the target paper list from Llama results.

    run_all=True  → all 470 papers Llama processed (full validation run)
    run_all=False → only Llama-excluded papers not in the 120-paper human set
    """
    # Load Llama
    wb = openpyxl.load_workbook(llama_path)
    ws = wb.active
    hdrs = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    llama = {}
    for r in range(2, ws.max_row + 1):
        row = {hdrs[c-1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
        if row.get("_filename"):
            llama[row["_filename"]] = row
    # Load 120-paper human-annotated set (for reporting only)
    wb2 = openpyxl.load_workbook(qwen120_path)
    ws2 = wb2.active
    hdrs2 = [ws2.cell(1, c).value for c in range(1, ws2.max_column + 1)]
    annotated = set()
    for r in range(2, ws2.max_row + 1):
        row = {hdrs2[c-1]: ws2.cell(r, c).value for c in range(1, ws2.max_column + 1)}
        if row.get("_filename"):
            annotated.add(row["_filename"])

    if run_all:
        target = sorted(llama.keys())
        print(f"  Mode: ALL papers  |  Llama total: {len(llama)}")
        print(f"  (Of these, {len(annotated)} are in the 120-paper human set — "
              f"results will be comparable)")
    else:
        target = sorted([
            p for p, v in llama.items()
            if v.get("include") is not True and p not in annotated
        ])
        print(f"  Mode: Llama-excludes only  |  Llama total: {len(llama)}  "
              f"|  Already annotated: {len(annotated)}")
        print(f"  Target (Llama-excluded, not annotated): {len(target)}")

    return target, llama


def find_pdf(filename, pdf_folder):
    """Locate a PDF by filename, searching folder and one level down."""
    direct = os.path.join(pdf_folder, filename)
    if os.path.exists(direct):
        return direct
    for root, _, files in os.walk(pdf_folder):
        if filename in files:
            return os.path.join(root, filename)
    return None


def extract_pdf_text(pdf_path, max_chars=MAX_TOKENS_INPUT * 4):
    """Extract text from PDF, preferring structured text files if present."""
    # Try zip-based text extraction first (the format used in this project)
    import zipfile, io
    try:
        with zipfile.ZipFile(pdf_path) as z:
            names = z.namelist()
            txt_files = sorted([n for n in names if n.endswith('.txt')
                                 and not n.startswith('__')])
            if txt_files:
                text = ""
                for fn in txt_files:
                    text += z.read(fn).decode('utf-8', errors='replace') + "\n"
                    if len(text) > max_chars:
                        break
                return text[:max_chars].strip()
    except Exception:
        pass
    # Fall back to PyPDF2
    try:
        import PyPDF2
        with open(pdf_path, 'rb') as f:
            reader = PyPDF2.PdfReader(f)
            text = ""
            for page in reader.pages:
                page_text = page.extract_text() or ""
                text += page_text + "\n"
                if len(text) > max_chars:
                    break
        return text[:max_chars].strip()
    except Exception as e:
        print(f"    PDF read error: {e}")
        return None


def parse_json(text):
    """Extract and parse the first JSON object from model output."""
    # Strip markdown fences
    text = re.sub(r'```json\s*', '', text)
    text = re.sub(r'```\s*', '', text)
    match = re.search(r'\{[\s\S]*\}', text)
    if not match:
        return None
    try:
        raw = match.group()
        # Fix common model JSON errors
        raw = re.sub(r',\s*}', '}', raw)
        raw = re.sub(r',\s*]', ']', raw)
        raw = re.sub(r'true_or_false', 'null', raw)
        raw = re.sub(r'1_to_5', '3', raw)
        raw = re.sub(r'true_or_false_or_null', 'null', raw)
        return json.loads(raw)
    except json.JSONDecodeError:
        return None


def flatten_result(filename, parsed, llama_row, raw_response):
    """Flatten the parsed JSON into a single dict suitable for a spreadsheet row."""
    if parsed is None:
        return {
            "_filename": filename,
            "_parse_error": True,
            "_raw_response": str(raw_response)[:500],
            "llama_include": llama_row.get("include") if llama_row else None,
            "llama_confidence": llama_row.get("confidence") if llama_row else None,
            "llama_rationale": str(llama_row.get("decision_rationale",""))[:200] if llama_row else "",
        }

    c4 = parsed.get("C4_study_type", {})
    c1 = parsed.get("C1_population", {})
    c2 = parsed.get("C2_ai_technology", {})
    c3 = parsed.get("C3_communication_support", {})
    sd = parsed.get("screening_decision", {})

    # ── Derive include from criterion fields (don't trust model's own include) ──
    # This prevents the model from including papers while simultaneously
    # reporting that criteria are not met (a common LLM inconsistency).

    is_review  = c4.get("is_review_article", False)
    c2_tier_raw = c2.get("tier", "none")          # model's original claim
    c2_tier     = c2_tier_raw                     # may be corrected by whitelist below

    # Post-processing: validate Tier A claims against known non-AI term list.
    # If the model listed only interface/design terms as "explicit AI", downgrade
    # the tier to none and force exclusion on C2.
    # Whitelist validation for Tier A: require at least one genuine AI term.
    # The model often lists interface/design terms (touchscreen, Skype, hypermedia)
    # as "explicit AI". If none of the listed terms match the AI whitelist,
    # downgrade to tier=none regardless of the model's claim.
    REAL_AI_TERMS = {
        # Core AI/ML vocabulary
        "artificial intelligence","machine learning","deep learning","neural network",
        "neural net","reinforcement learning","supervised learning","transformer",
        # Language models
        "large language model","llm","gpt","gpt-4","gpt-3","chatgpt","bert","t5",
        "language model","generative ai","generative model",
        # NLP
        "natural language processing","nlp","natural language understanding","nlu",
        "natural language generation","nlg","text generation","sentiment analysis",
        "named entity recognition","intent recognition","intent classification",
        "dialogue system","dialogue management","dialogue manager",
        # Speech AI (must be understanding/generation, not just playback)
        "automatic speech recognition","asr","speech recognition",
        "speech synthesis","text-to-speech","tts","voice synthesis",
        "speech understanding","spoken language understanding",
        # Conversational AI
        "conversational agent","conversational ai","conversational system",
        "chatbot","virtual agent","embodied conversational agent","eca",
        "voice assistant","intelligent assistant","dialogue agent",
        # Computer vision / multimodal AI
        "computer vision","image recognition","facial recognition",
        "emotion recognition","affect recognition","gesture recognition",
        "object detection","object recognition",
        # Word/text prediction
        "word prediction","next-word prediction","predictive text",
        "augmentative and alternative communication with ai",
        # Adaptive/personalised systems (must describe the mechanism)
        "adaptive algorithm","personalisation algorithm","recommendation system",
        "user modelling","user model",
        # Robotics with explicit AI
        "socially assistive robot with ai","robot with natural language",
        "autonomous robot","robot learning",
    }
    if c2_tier == "A_explicit":
        terms_raw = c2.get("explicit_ai_terms") or []
        if isinstance(terms_raw, str):
            terms_raw = [t.strip() for t in terms_raw.split(",") if t.strip()]
        evidence_text = str(c2.get("ai_evidence_quote","")).lower()
        # Check if any listed term OR the evidence text contains a real AI term
        all_text = " ".join(t.lower() for t in terms_raw) + " " + evidence_text
        has_real_ai = any(ai_term in all_text for ai_term in REAL_AI_TERMS)
        if not has_real_ai:
            c2_tier = "none"   # downgrade: no genuine AI term found
        # Teleoperated / wizard-of-oz → operator-controlled, not AI
        if "teleoperat" in evidence_text or "wizard of oz" in evidence_text:
            c2_tier = "none"

    c1_ok      = (c1.get("has_dementia") or c1.get("has_aphasia") or c1.get("has_ppa"))
    c2_ok      = c2_tier in ("A_explicit", "B_implicit")
    c2_basic   = c2.get("is_basic_digital_tool", False)
    c2_bg_only = c2.get("is_ai_only_in_background", False)
    c3_ok      = c3.get("is_direct_communication_support") is True
    c3_therapy = c3.get("is_therapy_training_app", False)
    c3_detect  = c3.get("is_detection_assessment_only", False)

    # Hard exclusion rules (override model's include=True)
    exclude_c4 = is_review
    exclude_c2 = (c2_tier == "none") or c2_basic or c2_bg_only
    exclude_c3 = c3_therapy or c3_detect or (not c3_ok)
    exclude_c1 = not c1_ok

    include = (not exclude_c4 and not exclude_c2 and not exclude_c3 and not exclude_c1)

    # Flag for human review: Tier B AI (implicit, needs verification) OR
    # model's own flag OR any criterion confidence <= 2 on a borderline paper
    flag = False
    if c2_tier == "B_implicit":
        flag = True  # Tier B always needs human check
    if sd.get("flag_for_human_review"):
        flag = True
    any_low_conf = any(
        c.get("confidence") is not None and c.get("confidence") <= 2
        for c in [c1, c2, c3]
    )
    if include and any_low_conf:
        flag = True
    if is_review:
        include = False
        flag    = False

    return {
        # Identification
        "_filename":             filename,
        "_parse_error":          False,

        # Final decision
        "include":               include,
        "flag_for_human_review": flag,
        "exclude_reason":        ",".join(filter(None, [
                                     "C4" if exclude_c4 else "",
                                     "C1" if exclude_c1 else "",
                                     "C2" if exclude_c2 else "",
                                     "C2_downgraded" if (exclude_c2 and c2_tier_raw == "A_explicit") else "",
                                     "C3" if exclude_c3 else "",
                                 ])) or "none",
        "overall_confidence":    sd.get("overall_confidence", ""),
        "decision_rationale":    sd.get("decision_rationale", ""),

        # C4 – Study type
        "c4_is_review":          c4.get("is_review_article", ""),
        "c4_study_type":         c4.get("study_type", ""),
        "c4_evidence":           c4.get("evidence", ""),

        # C1 – Population
        "c1_has_dementia":       c1.get("has_dementia", ""),
        "c1_has_aphasia":        c1.get("has_aphasia", ""),
        "c1_has_ppa":            c1.get("has_ppa", ""),
        "c1_confirmed":          c1.get("diagnosis_confirmed", ""),
        "c1_population_notes":   str(c1.get("population_notes", ""))[:300],
        "c1_confidence":         c1.get("confidence", ""),

        # C2 – AI (c2_tier = post-processed validated tier; c2_tier_raw = model's claim)
        "c2_tier":               c2_tier,
        "c2_tier_raw":           c2_tier_raw,
        "c2_explicit_terms":     ", ".join(c2.get("explicit_ai_terms", [])),
        "c2_implicit_indicators":  ", ".join(c2.get("implicit_ai_indicators", [])),
        "c2_technology_name":    c2.get("technology_name", ""),
        "c2_ai_evidence":        str(c2.get("ai_evidence_quote", ""))[:300],
        "c2_is_basic_tool":      c2.get("is_basic_digital_tool", ""),
        "c2_ai_only_background": c2.get("is_ai_only_in_background", ""),
        "c2_confidence":         c2.get("confidence", ""),

        # C3 – Communication
        "c3_direct_comm":        c3.get("is_direct_communication_support", ""),
        "c3_therapy_app":        c3.get("is_therapy_training_app", ""),
        "c3_detection_only":     c3.get("is_detection_assessment_only", ""),
        "c3_indirect_psych":     c3.get("is_indirect_psychosocial_only", ""),
        "c3_cognitive_only":     c3.get("is_cognitive_training_only", ""),
        "c3_comm_role":          str(c3.get("communication_role", ""))[:300],
        "c3_confidence":         c3.get("confidence", ""),

        # Llama round-1 context
        "llama_include":         llama_row.get("include") if llama_row else None,
        "llama_confidence":      llama_row.get("confidence") if llama_row else None,
        "llama_rationale":       str(llama_row.get("decision_rationale",""))[:200] if llama_row else "",
        "llama_has_ai":          llama_row.get("has_ai") if llama_row else None,
        "llama_svd":             llama_row.get("support_vs_detect") if llama_row else None,

        # Debug
        "_raw_response":         str(raw_response)[:300],
    }


def save_checkpoint(results, checkpoint_path):
    with open(checkpoint_path, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, default=str)


def load_checkpoint(checkpoint_path):
    if os.path.exists(checkpoint_path):
        with open(checkpoint_path, encoding='utf-8') as f:
            return json.load(f)
    return []


def write_xlsx(rows, filepath):
    """Write results to Excel with colour coding."""
    if not rows:
        print("  No rows to write.")
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Screening"
    headers = list(rows[0].keys())
    bold = Font(bold=True)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, h)
        c.font = bold
    fills = {
        "include_true":  PatternFill("solid", fgColor="C6EFCE"),
        "include_false": PatternFill("solid", fgColor="FFC7CE"),
        "flag":          PatternFill("solid", fgColor="FFEB9C"),
        "error":         PatternFill("solid", fgColor="DDDDDD"),
    }
    for ri, row in enumerate(rows, 2):
        if row.get("_parse_error"):
            fill = fills["error"]
        elif row.get("flag_for_human_review"):
            fill = fills["flag"]
        elif row.get("include"):
            fill = fills["include_true"]
        else:
            fill = fills["include_false"]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(ri, ci, row.get(h, ""))
            cell.fill = fill
    for ci, h in enumerate(headers, 1):
        max_len = max(len(str(h)),
                      max((len(str(r.get(h) or "")) for r in rows), default=0))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 60)
    wb.save(filepath)
    print(f"  Saved: {filepath}  ({len(rows)} rows)")


# =============================================================================
# MODEL LOADING
# =============================================================================

def load_model():
    print("\n[2] Loading model…")
    try:
        import torch
        from transformers import AutoModelForCausalLM, AutoTokenizer, BitsAndBytesConfig
    except ImportError:
        import subprocess
        subprocess.run([
            "pip", "install", "-q", "transformers", "torch", "bitsandbytes",
            "accelerate", "PyPDF2", "pandas", "openpyxl", "--break-system-packages"
        ])
        import torch
        from transformers import AutoModelForCausalLM, AutoTokenizer, BitsAndBytesConfig

    kwargs = {"token": HF_TOKEN} if HF_TOKEN else {}

    # Reduce memory fragmentation (suggested in PyTorch OOM messages)
    os.environ["PYTORCH_ALLOC_CONF"] = "expandable_segments:True"

    # ── Diagnose GPU state before loading ─────────────────────────────────────
    if torch.cuda.is_available():
        n_gpus = torch.cuda.device_count()
        print(f"  GPUs available: {n_gpus}")
        for i in range(n_gpus):
            total  = torch.cuda.get_device_properties(i).total_memory / 1e9
            reserved = torch.cuda.memory_reserved(i) / 1e9
            allocated = torch.cuda.memory_allocated(i) / 1e9
            free   = total - reserved
            print(f"  GPU {i}: {torch.cuda.get_device_name(i)}")
            print(f"    Total: {total:.1f} GB  |  Reserved: {reserved:.1f} GB  "
                  f"|  Allocated: {allocated:.1f} GB  |  Free: {free:.1f} GB")
        # Clear any leftover cache from previous runs / crashes
        torch.cuda.empty_cache()
        gc.collect()
    else:
        print("  No GPU found.")

    tokenizer = AutoTokenizer.from_pretrained(MODEL_NAME, **kwargs)

    # ── Load model ─────────────────────────────────────────────────────────────
    # Try 4-bit quantization first. If bitsandbytes is unavailable or errors,
    # fall back to fp16 (needs ~14 GB; fine on 46 GB GPU).
    if torch.cuda.is_available():
        print(f"  Loading in 4-bit quantization…")
        try:
            bnb_config = BitsAndBytesConfig(
                load_in_4bit=True,
                bnb_4bit_use_double_quant=True,
                bnb_4bit_quant_type="nf4",
                bnb_4bit_compute_dtype=torch.bfloat16,
            )
            model = AutoModelForCausalLM.from_pretrained(
                MODEL_NAME, quantization_config=bnb_config,
                device_map="auto", **kwargs
            )
            print("  4-bit load: OK")
        except Exception as e:
            print(f"  4-bit load failed ({e}); falling back to fp16…")
            model = AutoModelForCausalLM.from_pretrained(
                MODEL_NAME, torch_dtype=torch.float16,
                device_map="auto", **kwargs
            )
            print("  fp16 load: OK")
        # Report actual memory after loading
        for i in range(torch.cuda.device_count()):
            alloc = torch.cuda.memory_allocated(i) / 1e9
            print(f"  GPU {i} after load: {alloc:.1f} GB allocated")
    else:
        print("  No GPU – loading on CPU (slow)")
        model = AutoModelForCausalLM.from_pretrained(
            MODEL_NAME, torch_dtype="auto", device_map="cpu", **kwargs
        )

    model.eval()
    print(f"  Model ready: {MODEL_NAME}")
    return model, tokenizer


def run_model(model, tokenizer, prompt_text):
    """Run the model on a single prompt and return the raw text response."""
    import torch
    messages = [
        {
            "role": "system",
            "content": (
                "You are a research assistant screening academic articles for a "
                "systematic review. You MUST return ONLY valid JSON with no text "
                "before or after. Follow the JSON schema exactly."
            )
        },
        {"role": "user", "content": prompt_text}
    ]
    formatted = tokenizer.apply_chat_template(
        messages, tokenize=False, add_generation_prompt=True
    )
    inputs = tokenizer(formatted, return_tensors="pt").to(model.device)
    with torch.no_grad():
        outputs = model.generate(
            **inputs,
            max_new_tokens=MAX_TOKENS_OUTPUT,
            temperature=TEMPERATURE,
            top_p=0.9,
            repetition_penalty=1.05,
            do_sample=True,
            pad_token_id=tokenizer.eos_token_id,
        )
    response = tokenizer.decode(
        outputs[0][inputs["input_ids"].shape[1]:], skip_special_tokens=True
    )
    del inputs, outputs
    gc.collect()
    if hasattr(model, 'device') and str(model.device) != 'cpu':
        import torch; torch.cuda.empty_cache()
    return response


# =============================================================================
# MAIN
# =============================================================================

def main():
    print("=" * 68)
    print("  Enabling Voices – Qwen Screening v3")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("=" * 68)

    # ── [1] Build paper list ──────────────────────────────────────────────────
    print("\n[1] Building target paper list…")
    if PAPER_OVERRIDE:
        papers = PAPER_OVERRIDE
        llama_data = {}
        print(f"  Using PAPER_OVERRIDE: {len(papers)} papers")
    else:
        papers, llama_data = load_paper_list(LLAMA_FILE, QWEN_120_FILE,
                                             run_all=RUN_ALL_PAPERS)

    if TEST_MODE_LIMIT:
        papers = papers[:TEST_MODE_LIMIT]
        print(f"  TEST MODE: limiting to {TEST_MODE_LIMIT} papers")

    total = len(papers)
    print(f"  Papers to screen: {total}")

    # ── [2] Load model ────────────────────────────────────────────────────────
    model, tokenizer = load_model()

    # ── [3] Set up output ─────────────────────────────────────────────────────
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    timestamp    = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file  = f"{OUTPUT_DIR}/{OUTPUT_PREFIX}_{timestamp}.xlsx"
    checkpoint_f = f"{OUTPUT_DIR}/{OUTPUT_PREFIX}_checkpoint.json"

    # Resume from checkpoint if present
    done_results = load_checkpoint(checkpoint_f)
    done_files   = {r["_filename"] for r in done_results}
    remaining    = [p for p in papers if p not in done_files]
    print(f"  Already done (checkpoint): {len(done_results)}  |  Remaining: {len(remaining)}")

    results = list(done_results)

    # ── [4] Process papers ────────────────────────────────────────────────────
    print("\n[3] Processing papers…\n")
    for i, filename in enumerate(remaining, 1):
        overall_idx = len(done_results) + i
        print(f"  [{overall_idx}/{total}] {filename}")

        # Find PDF
        pdf_path = find_pdf(filename, PDF_FOLDER)
        if not pdf_path:
            print(f"    SKIP – PDF not found")
            results.append({
                "_filename": filename,
                "_parse_error": True,
                "decision_rationale": "PDF not found",
                "_raw_response": "file_not_found",
                "include": False,
                "flag_for_human_review": False,
            })
            continue

        # Extract text
        text = extract_pdf_text(pdf_path)
        if not text or len(text) < 100:
            print(f"    SKIP – insufficient text extracted ({len(text) if text else 0} chars)")
            results.append({
                "_filename": filename,
                "_parse_error": True,
                "decision_rationale": "Insufficient text",
                "_raw_response": "text_extraction_failed",
                "include": False,
                "flag_for_human_review": True,
            })
            continue

        print(f"    Text: {len(text)} chars  |  PDF: {pdf_path}")

        # Run model
        prompt = SCREENING_PROMPT.format(article_text=text)
        try:
            raw = run_model(model, tokenizer, prompt)
        except Exception as e:
            print(f"    MODEL ERROR: {e}")
            results.append({
                "_filename": filename,
                "_parse_error": True,
                "decision_rationale": f"Model error: {str(e)[:100]}",
                "_raw_response": str(e)[:300],
                "include": False,
                "flag_for_human_review": True,
            })
            continue

        # Parse JSON
        parsed = parse_json(raw)
        if parsed is None:
            print(f"    PARSE ERROR – raw: {raw[:150]}")
        else:
            sd = parsed.get("screening_decision", {})
            print(f"    include={sd.get('include')}  conf={sd.get('overall_confidence')}  "
                  f"tier={parsed.get('C2_ai_technology',{}).get('tier','')}  "
                  f"flag={sd.get('flag_for_human_review')}")
            print(f"    {sd.get('decision_rationale','')}")

        llama_row = llama_data.get(filename)
        row = flatten_result(filename, parsed, llama_row, raw)
        results.append(row)

        # Checkpoint
        if i % CHECKPOINT_EVERY == 0:
            save_checkpoint(results, checkpoint_f)
            write_xlsx(results, output_file.replace(".xlsx", "_partial.xlsx"))
            print(f"    ── checkpoint saved ({len(results)} papers) ──")

    # ── [5] Final output ──────────────────────────────────────────────────────
    save_checkpoint(results, checkpoint_f)
    write_xlsx(results, output_file)

    # Summary statistics
    includes   = [r for r in results if r.get("include")]
    flags      = [r for r in results if r.get("flag_for_human_review")]
    errors     = [r for r in results if r.get("_parse_error")]
    tier_a     = [r for r in results if r.get("c2_tier") == "A_explicit"]
    tier_b     = [r for r in results if r.get("c2_tier") == "B_implicit"]

    print("\n" + "=" * 68)
    print("SCREENING COMPLETE")
    print(f"  Total processed:      {len(results)}")
    print(f"  Include:              {len(includes)}")
    print(f"  Flag for human review:{len(flags)}")
    print(f"  Errors/skipped:       {len(errors)}")
    print(f"  Tier A (explicit AI): {len(tier_a)}")
    print(f"  Tier B (implicit AI): {len(tier_b)}")
    print(f"\n  Output: {output_file}")
    print("=" * 68)

    # Write a separate flagged-only file for ease of review
    flag_file = output_file.replace(".xlsx", "_flagged.xlsx")
    write_xlsx(flags, flag_file)
    inc_file  = output_file.replace(".xlsx", "_includes.xlsx")
    write_xlsx(includes, inc_file)

    return output_file


if __name__ == "__main__":
    main()
